"""Script to create example Excel file for testing.

Run this script after installing dependencies:
    pip install -r requirements.txt

This will create examples/test_financial_statement.xlsx with sample financial data.
"""
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install dependencies first: pip install -r requirements.txt")
    sys.exit(1)

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Financial Statement"

# Set header row
headers = ["item_name", "category", "amount", "notes"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Sample data rows
sample_data = [
    ("Cash at bank", "Cash", 1200000, "Bank account balance"),
    ("Finished goods", "Inventory", 800000, "Ready for sale"),
    ("Customer debts", "Receivable", 500000, "Trade receivables"),
    ("Supplier payables", "Liability", 300000, "Short-term payables"),
    ("Share capital", "Capital", 10000000, "Paid-in capital"),
    ("Retained earnings", "Retained Earnings", 2000000, "Accumulated profits"),
]

# Add data rows
for row_idx, (item_name, category, amount, notes) in enumerate(sample_data, start=2):
    ws.cell(row=row_idx, column=1, value=item_name)
    ws.cell(row=row_idx, column=2, value=category)
    ws.cell(row=row_idx, column=3, value=amount)
    ws.cell(row=row_idx, column=4, value=notes)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the file
import os
output_path = "examples/test_financial_statement.xlsx"
os.makedirs("examples", exist_ok=True)
wb.save(output_path)
print(f"✓ Created example Excel file: {output_path}")
print(f"  File contains {len(sample_data)} sample financial statement items.")
