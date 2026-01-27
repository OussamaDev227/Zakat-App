"""Excel file parser for financial statement data."""
from typing import List, Dict, Tuple, Any, Optional
from decimal import Decimal, InvalidOperation
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse Excel file and extract financial data rows.
    
    Args:
        file_path: Path to .xlsx file
        
    Returns:
        List of dictionaries, each representing a row with keys:
        - item_name (str)
        - category (str)
        - amount (Decimal or None if invalid)
        - notes (str or None)
        - row_number (int): Excel row number (1-indexed)
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not a valid Excel file
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active  # Use first sheet
        
        rows = []
        header_row = None
        header_map = {}  # Maps column letter to field name
        
        # Find header row (first non-empty row)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=False), start=1):
            values = [cell.value for cell in row if cell.value is not None]
            if values:
                # Check if this looks like a header row
                header_candidates = [str(v).lower().strip() for v in values]
                if any(keyword in ' '.join(header_candidates) for keyword in ['item', 'category', 'amount', 'name']):
                    header_row = row_idx
                    # Map columns to field names
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value:
                            col_letter = get_column_letter(col_idx)
                            value_lower = str(cell.value).lower().strip()
                            # Map common column name variations
                            if 'item' in value_lower and 'name' in value_lower:
                                header_map[col_letter] = 'item_name'
                            elif 'name' in value_lower and 'item' not in value_lower:
                                header_map[col_letter] = 'item_name'
                            elif 'category' in value_lower:
                                header_map[col_letter] = 'category'
                            elif 'amount' in value_lower or 'value' in value_lower:
                                header_map[col_letter] = 'amount'
                            elif 'note' in value_lower or 'comment' in value_lower or 'description' in value_lower:
                                header_map[col_letter] = 'notes'
                    break
        
        if not header_row:
            # Fallback: assume first row is header with standard column order
            # item_name, category, amount, notes
            header_map = {'A': 'item_name', 'B': 'category', 'C': 'amount', 'D': 'notes'}
            header_row = 1
        
        # Parse data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            row_data = {}
            has_data = False
            
            for col_letter, field_name in header_map.items():
                col_idx = column_index_from_string(col_letter)
                if col_idx <= len(row):
                    cell = row[col_idx - 1]
                    value = cell.value
                    
                    if field_name == 'amount':
                        # Try to parse as number
                        if value is None:
                            row_data[field_name] = None
                        elif isinstance(value, (int, float)):
                            try:
                                row_data[field_name] = Decimal(str(value))
                            except (InvalidOperation, ValueError):
                                row_data[field_name] = None
                        else:
                            # Try to parse string as number
                            try:
                                # Remove commas and whitespace
                                clean_value = str(value).replace(',', '').replace(' ', '').strip()
                                row_data[field_name] = Decimal(clean_value)
                            except (InvalidOperation, ValueError):
                                row_data[field_name] = None
                    else:
                        # String fields
                        if value is not None:
                            row_data[field_name] = str(value).strip()
                            if row_data[field_name]:
                                has_data = True
                        else:
                            row_data[field_name] = None
                else:
                    row_data[field_name] = None
            
            # Only include row if it has at least item_name or category
            if has_data or row_data.get('item_name') or row_data.get('category'):
                row_data['row_number'] = row_idx
                rows.append(row_data)
        
        workbook.close()
        return rows
        
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")


def validate_excel_structure(rows: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
    """
    Validate that parsed rows have required structure.
    
    Args:
        rows: List of parsed row dictionaries
        
    Returns:
        Tuple of (is_valid: bool, errors: List[str])
    """
    errors = []
    
    if not rows:
        errors.append("Excel file contains no data rows")
        return False, errors
    
    # Check for required fields in at least some rows
    has_item_name = any(row.get('item_name') for row in rows)
    has_category = any(row.get('category') for row in rows)
    has_amount = any(row.get('amount') is not None for row in rows)
    
    if not has_item_name:
        errors.append("Missing required column: item_name")
    if not has_category:
        errors.append("Missing required column: category")
    if not has_amount:
        errors.append("Missing required column: amount")
    
    is_valid = has_item_name and has_category and has_amount
    return is_valid, errors
